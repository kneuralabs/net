import os
import io
import time
import base64
import logging
import warnings
import bcrypt
import openpyxl
import requests
from threading import Lock
from urllib.parse import urlparse
from cryptography.fernet import Fernet
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from itsdangerous import URLSafeTimedSerializer
from flask import Flask, render_template, request, redirect, flash

logger = logging.getLogger(__name__)

app = Flask(__name__)

# ── Secrets ────────────────────────────────────────────────────────────────────
# Production (the default) requires all secrets via the environment and fails
# fast otherwise. Dev mode is an explicit opt-in (NET_DEV=1 / FLASK_DEBUG /
# FLASK_ENV=development); legacy committed defaults additionally require
# NET_ALLOW_INSECURE_DEFAULTS=1 and emit a loud warning.

def _is_dev_mode():
    return (
        os.environ.get('NET_DEV', '') == '1'
        or os.environ.get('FLASK_DEBUG', '').lower() in ('1', 'true')
        or os.environ.get('FLASK_ENV', '').lower() == 'development'
    )


def _allow_insecure_defaults():
    return os.environ.get('NET_ALLOW_INSECURE_DEFAULTS', '') == '1'


if not _is_dev_mode():
    _missing = [name for name in ('SSO_SHARED_SECRET', 'EXCEL_PASSWORD')
                if not os.environ.get(name)]
    if _missing:
        raise RuntimeError(
            'Refusing to start in production without required secrets. '
            'Missing environment variables: ' + ', '.join(_missing) + '. '
            'Set them, or set NET_DEV=1 for local development.'
        )

_app_secret = os.environ.get('SSO_APP_SECRET', '')
if not _app_secret:
    warnings.warn(
        'SSO_APP_SECRET is not set — using a random per-process key. '
        'Flash messages/sessions will not survive a restart.',
        RuntimeWarning, stacklevel=1,
    )
    _app_secret = 'insecure-dev-' + os.urandom(32).hex()
app.secret_key = _app_secret

# Session-cookie hardening. Secure is relaxed only in explicit dev mode,
# where the app is served over plain HTTP.
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SECURE=not _is_dev_mode(),
    SESSION_COOKIE_SAMESITE='Lax',
    MAX_CONTENT_LENGTH=1 * 1024 * 1024,  # request bodies are small forms only
)


@app.after_request
def _security_headers(resp):
    resp.headers.setdefault('X-Content-Type-Options', 'nosniff')
    resp.headers.setdefault('X-Frame-Options', 'DENY')
    resp.headers.setdefault('Referrer-Policy', 'strict-origin-when-cross-origin')
    if not _is_dev_mode():
        resp.headers.setdefault('Strict-Transport-Security',
                                'max-age=31536000; includeSubDomains')
    return resp


SSO_SHARED_SECRET = os.environ.get('SSO_SHARED_SECRET', '')
if not SSO_SHARED_SECRET:
    # Dev mode only (production raised above).
    warnings.warn(
        'SSO_SHARED_SECRET is not set — issued tokens cannot be verified by '
        'the intranet. Set the SSO_SHARED_SECRET environment variable.',
        RuntimeWarning, stacklevel=1,
    )
    SSO_SHARED_SECRET = 'insecure-dev-' + os.urandom(32).hex()

EXCEL_PASSWORD = os.environ.get('EXCEL_PASSWORD', '')
if not EXCEL_PASSWORD:
    if _allow_insecure_defaults():
        warnings.warn(
            'NET_ALLOW_INSECURE_DEFAULTS=1: using the legacy INSECURE dev '
            'EXCEL_PASSWORD. Never use this outside local development.',
            RuntimeWarning, stacklevel=1,
        )
        EXCEL_PASSWORD = 'KneuraExcel@2026'
    else:
        raise RuntimeError(
            'EXCEL_PASSWORD is not set. It is required to open DATA.xlsx. '
            'Set the EXCEL_PASSWORD environment variable (or, for local dev '
            'only, also set NET_ALLOW_INSECURE_DEFAULTS=1).'
        )

EMPLOYEE_DATA_URL = 'https://raw.githubusercontent.com/kneuralabs/ID/main/EmployeeData.xlsx'

# Initial password accepted for first-time users (changed on first login).
DEFAULT_PASSWORD = os.environ.get('DEFAULT_USER_PASSWORD', '')
if not DEFAULT_PASSWORD:
    if _allow_insecure_defaults():
        warnings.warn(
            'NET_ALLOW_INSECURE_DEFAULTS=1: using the legacy INSECURE default '
            'user onboarding password. Set DEFAULT_USER_PASSWORD instead.',
            RuntimeWarning, stacklevel=1,
        )
        DEFAULT_PASSWORD = 'Kneuralabs@2026'
    else:
        raise RuntimeError(
            'DEFAULT_USER_PASSWORD is not set. It is required for first-login '
            'onboarding. Set the DEFAULT_USER_PASSWORD environment variable '
            '(or, for local dev only, NET_ALLOW_INSECURE_DEFAULTS=1).'
        )

# DATA.xlsx is shared with the intranet; configure the path via env var
DATA_FILE = os.environ.get(
    'DATA_FILE',
    os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'DATA.xlsx')
)

def _parse_hosts(raw):
    """Parse a comma-separated host allow-list into a normalised set."""
    return set(h.strip().lower() for h in raw.split(',') if h.strip())


def _host_allowed(netloc, allowed):
    """True when a URL's netloc matches an allow-list entry.

    Entries are either exact hosts ("intranet.kneuralabs.com") or a
    wildcard ("*.kneuralabs.com") that matches the apex domain and any
    subdomain depth. Any userinfo and port in netloc are ignored. Wildcard
    matching is anchored on a leading dot, so look-alikes such as
    "kneuralabs.com.evil.com" and "evilkneuralabs.com" are rejected.
    """
    host = netloc.split('@')[-1].split(':')[0].strip('.').lower()
    if not host:
        return False
    for entry in allowed:
        if entry.startswith('*.'):
            base = entry[2:]
            if host == base or host.endswith('.' + base):
                return True
        elif host == entry:
            return True
    return False


# Hosts allowed for token-bearing callbacks (server-side intranet session).
_ALLOWED_HOSTS = _parse_hosts(
    os.environ.get('ALLOWED_CALLBACK_HOSTS', 'intranet.kneuralabs.com')
)

# Hosts allowed for direct (no-token) post-login redirects.
# Every workspace widget now lives on a *.kneuralabs.com subdomain, so the
# allow-list tracks the apex + all subdomains rather than a fixed list that
# silently breaks each time a tool is added or its URL changes. The legacy
# kneuralabs.github.io host is kept for backward compatibility.
_ALLOWED_REDIRECT_HOSTS = _parse_hosts(
    os.environ.get('ALLOWED_REDIRECT_HOSTS',
                   '*.kneuralabs.com,kneuralabs.com,kneuralabs.github.io')
)


def _is_valid_redirect(url):
    """Allow HTTPS redirects to whitelisted widget hosts."""
    try:
        p = urlparse(url)
        return p.scheme == 'https' and _host_allowed(p.netloc, _ALLOWED_REDIRECT_HOSTS)
    except Exception:
        return False

_SALT = b'kneura_labs_2026'


def _fernet():
    kdf = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=_SALT, iterations=390000)
    return Fernet(base64.urlsafe_b64encode(kdf.derive(EXCEL_PASSWORD.encode())))


def _load_wb():
    if not os.path.exists(DATA_FILE) or os.path.getsize(DATA_FILE) == 0:
        return None
    try:
        with open(DATA_FILE, 'rb') as f:
            raw = f.read()
        decrypted = _fernet().decrypt(raw)
        return openpyxl.load_workbook(io.BytesIO(decrypted))
    except Exception:
        return None


def _get_user(employee_id):
    wb = _load_wb()
    if wb is None:
        return None
    for row in wb.active.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and str(row[0]).strip() == str(employee_id).strip():
            return {'id': row[0], 'password_hash': row[1], 'status': row[2]}
    return None


# ── Roster cache ───────────────────────────────────────────────────────────────
# The EmployeeData.xlsx roster is fetched from GitHub at most once per hour.
# On fetch failure we fall back to the last good copy (with a warning) so
# logins keep working through brief GitHub outages.
_ROSTER_TTL = 3600  # seconds
_roster_lock = Lock()
_roster_cache = {'rows': None, 'fetched_at': 0.0}


def _fetch_roster_rows():
    """Download the roster and return its data rows as a list of tuples."""
    resp = requests.get(EMPLOYEE_DATA_URL, timeout=10)
    resp.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    return list(wb.active.iter_rows(min_row=2, values_only=True))


def _get_roster_rows():
    """Return cached roster rows, refreshing at most every _ROSTER_TTL seconds.

    Returns None only when no fetch has ever succeeded."""
    with _roster_lock:
        now = time.time()
        fresh = (_roster_cache['rows'] is not None
                 and now - _roster_cache['fetched_at'] < _ROSTER_TTL)
        if not fresh:
            try:
                _roster_cache['rows'] = _fetch_roster_rows()
                _roster_cache['fetched_at'] = now
            except Exception as exc:
                if _roster_cache['rows'] is not None:
                    logger.warning(
                        'Roster refresh failed (%s); falling back to cached '
                        'copy from %.0f seconds ago.',
                        exc, now - _roster_cache['fetched_at'],
                    )
                else:
                    logger.warning('Roster fetch failed and no cached copy '
                                   'is available: %s', exc)
        return _roster_cache['rows']


def _check_roster(employee_id):
    rows = _get_roster_rows()
    if rows is None:
        return 'error'
    revoke_kw = {'revok', 'inactive', 'terminat', 'disabled', 'suspended'}
    for row in rows:
        if not row or row[0] is None:
            continue
        if str(row[0]).strip() == str(employee_id).strip():
            for cell in row[1:]:
                if cell and any(k in str(cell).lower() for k in revoke_kw):
                    return 'revoked'
            return 'found'
    return 'not_found'


def _is_valid_callback(url):
    """Only allow HTTPS callbacks to whitelisted intranet hosts."""
    try:
        p = urlparse(url)
        return p.scheme == 'https' and _host_allowed(p.netloc, _ALLOWED_HOSTS)
    except Exception:
        return False


def _make_token(employee_id):
    s = URLSafeTimedSerializer(SSO_SHARED_SECRET)
    return s.dumps({'employee_id': str(employee_id)}, salt='sso-callback')


@app.route('/')
def index():
    callback = request.args.get('callback', '')
    redirect_url = request.args.get('redirect', '')
    if redirect_url:
        return redirect(f'/login?redirect={redirect_url}')
    return redirect(f'/login?callback={callback}' if callback else '/login')


@app.route('/login', methods=['GET', 'POST'])
def login():
    callback = request.args.get('callback', '').strip()
    redirect_url = request.args.get('redirect', '').strip()

    if request.method == 'POST':
        employee_id = request.form.get('employee_id', '').strip()
        password = request.form.get('password', '').strip()
        callback = request.form.get('callback', '').strip()
        redirect_url = request.form.get('redirect_url', '').strip()

        if not employee_id or not password:
            flash('Please enter both Employee ID and password.', 'error')
            return render_template('login.html', callback=callback)

        authenticated = False
        user = _get_user(employee_id)

        if user:
            if bcrypt.checkpw(password.encode(), user['password_hash'].encode()):
                authenticated = True
        else:
            # First-time user — check against external employee roster
            status = _check_roster(employee_id)
            if status == 'revoked':
                flash('Your account has been deactivated. Contact IT.', 'error')
                return render_template('login.html', callback=callback)
            if status == 'not_found':
                flash('Employee ID not found.', 'error')
                return render_template('login.html', callback=callback)
            if status == 'error':
                flash('Unable to verify employee. Please try again later.', 'error')
                return render_template('login.html', callback=callback)
            # status == 'found': accept the default password only
            if password == DEFAULT_PASSWORD:
                authenticated = True

        if not authenticated:
            flash('Invalid credentials.', 'error')
            return render_template('login.html', callback=callback)

        if redirect_url and _is_valid_redirect(redirect_url):
            return render_template('sso_success.html', redirect_url=redirect_url)

        if callback and _is_valid_callback(callback):
            token = _make_token(employee_id)
            sep = '&' if '?' in callback else '?'
            return redirect(f'{callback}{sep}token={token}')

        # No valid callback — show confirmation in place
        flash('Sign-in successful. Return to the intranet.', 'success')
        return render_template('login.html', callback=callback, redirect_url='')

    return render_template('login.html', callback=callback, redirect_url=redirect_url)


if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=5001)
