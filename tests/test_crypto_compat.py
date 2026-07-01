"""Regression tests for the intranet <-> SSO shared-store encryption contract.

Both ``app.py`` (the only writer) and ``sso/sso_app.py`` (a reader) derive a
Fernet key from the same ``EXCEL_PASSWORD`` to access the *same* ``DATA.xlsx``.
If their PBKDF2 salts drift apart the keys differ, every SSO-side decrypt raises
``InvalidToken``, ``_load_wb()`` swallows it and returns ``None``, and the SSO
server silently treats every employee as a first-time user — so only the shared
default password is ever accepted and per-user passwords stop being enforced.

These tests fail loudly if that ever regresses.
"""

import io

import openpyxl

import app
import sso_app


def test_kdf_salts_match():
    """The two salts must be byte-for-byte identical (cheap drift alarm)."""
    assert app._KDF_SALT == sso_app._SALT


def test_cross_app_decrypt_roundtrip():
    """A workbook encrypted by the intranet must decrypt on the SSO side."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['EmployeeID', 'PasswordHash', 'Status', 'LastLogin', 'CreatedAt'])
    ws.append(['E123', 'bcrypt-hash', 'active', '', '2026-01-01'])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    blob = app._fernet().encrypt(buf.read())
    # Raises InvalidToken if the keys diverge — exactly the shipped bug.
    decrypted = sso_app._fernet().decrypt(blob)

    reloaded = openpyxl.load_workbook(io.BytesIO(decrypted))
    first = list(reloaded.active.iter_rows(min_row=2, values_only=True))[0]
    assert first[0] == 'E123'


def test_sso_reads_user_written_by_intranet(tmp_path, monkeypatch):
    """End-to-end: a user the intranet writes is visible to the SSO server."""
    data_file = str(tmp_path / 'DATA.xlsx')
    monkeypatch.setattr(app, 'DATA_FILE', data_file)
    monkeypatch.setattr(app, '_wb_cache', {'wb': None, 'mtime': 0.0})
    monkeypatch.setattr(sso_app, 'DATA_FILE', data_file)

    app.upsert_user('E777', 'bcrypt-hash-value', status='active')

    user = sso_app._get_user('E777')
    assert user is not None
    assert str(user['id']) == 'E777'
    assert user['password_hash'] == 'bcrypt-hash-value'
    assert sso_app._get_user('E404') is None
