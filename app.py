import os
import io
import re
import time
import base64
import warnings
import bcrypt
import openpyxl
import requests
from datetime import datetime
from functools import wraps
from threading import Lock
from cryptography.fernet import Fernet
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify

app = Flask(__name__)

# ── Secrets ────────────────────────────────────────────────────────────────────
# In production (the default) all secrets MUST come from the environment and
# the app refuses to start otherwise. Dev mode must be opted into explicitly
# via NET_DEV=1 (or FLASK_DEBUG / FLASK_ENV=development).

def _is_dev_mode() -> bool:
    return (
        os.environ.get('NET_DEV', '') == '1'
        or os.environ.get('FLASK_DEBUG', '').lower() in ('1', 'true')
        or os.environ.get('FLASK_ENV', '').lower() == 'development'
    )


def _allow_insecure_defaults() -> bool:
    return os.environ.get('NET_ALLOW_INSECURE_DEFAULTS', '') == '1'


_DEV_MODE = _is_dev_mode()

if not _DEV_MODE:
    _missing = [name for name in ('SECRET_KEY', 'SSO_SHARED_SECRET', 'EXCEL_PASSWORD')
                if not os.environ.get(name)]
    if _missing:
        raise RuntimeError(
            'Refusing to start in production without required secrets. '
            'Missing environment variables: ' + ', '.join(_missing) + '. '
            'Set them, or set NET_DEV=1 for local development.'
        )

_SECRET_KEY = os.environ.get('SECRET_KEY', '')
if not _SECRET_KEY:
    # Dev mode only (production raised above): per-process random key.
    warnings.warn(
        'SECRET_KEY is not set — using a random per-process dev key. '
        'Sessions will not survive a restart. Set SECRET_KEY before deploying.',
        RuntimeWarning, stacklevel=1,
    )
    _SECRET_KEY = 'insecure-dev-' + os.urandom(32).hex()
app.secret_key = _SECRET_KEY

# Session-cookie hardening. Secure is relaxed only in explicit dev mode,
# where the app is served over plain HTTP.
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SECURE=not _DEV_MODE,
    SESSION_COOKIE_SAMESITE='Lax',
    MAX_CONTENT_LENGTH=1 * 1024 * 1024,  # request bodies are small forms only
)


@app.after_request
def _security_headers(resp):
    resp.headers.setdefault('X-Content-Type-Options', 'nosniff')
    resp.headers.setdefault('X-Frame-Options', 'DENY')
    resp.headers.setdefault('Referrer-Policy', 'strict-origin-when-cross-origin')
    if not _DEV_MODE:
        resp.headers.setdefault('Strict-Transport-Security',
                                'max-age=31536000; includeSubDomains')
    return resp


EXCEL_PASSWORD = os.environ.get('EXCEL_PASSWORD', '')
if not EXCEL_PASSWORD:
    # Dev mode only. The legacy committed password is gated behind an
    # explicit opt-in so it can never be a *silent* fallback.
    if _allow_insecure_defaults():
        warnings.warn(
            'NET_ALLOW_INSECURE_DEFAULTS=1: using the legacy INSECURE dev '
            'EXCEL_PASSWORD. Never use this outside local development.',
            RuntimeWarning, stacklevel=1,
        )
        EXCEL_PASSWORD = 'KneuraExcel@2026'
    else:
        raise RuntimeError(
            'EXCEL_PASSWORD is not set. It is required to open DATA.xlsx. '
            'Set the EXCEL_PASSWORD environment variable (or, for local dev '
            'only, also set NET_ALLOW_INSECURE_DEFAULTS=1 to use the legacy '
            'default).'
        )

DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'DATA.xlsx')
EMPLOYEE_DATA_URL = 'https://raw.githubusercontent.com/kneuralabs/ID/main/EmployeeData.xlsx'
# Initial password issued to first-time users (must be changed on first login).
DEFAULT_PASSWORD = os.environ.get('DEFAULT_USER_PASSWORD', '')
if not DEFAULT_PASSWORD:
    if _allow_insecure_defaults():
        warnings.warn(
            'NET_ALLOW_INSECURE_DEFAULTS=1: using the legacy INSECURE default '
            'user onboarding password. Set DEFAULT_USER_PASSWORD instead.',
            RuntimeWarning, stacklevel=1,
        )
        DEFAULT_PASSWORD = 'Kneuralabs@2026'
    else:
        raise RuntimeError(
            'DEFAULT_USER_PASSWORD is not set. It is required for first-login '
            'onboarding. Set the DEFAULT_USER_PASSWORD environment variable '
            '(or, for local dev only, NET_ALLOW_INSECURE_DEFAULTS=1 to use '
            'the legacy default).'
        )

# SSO configuration
SSO_URL            = os.environ.get('SSO_URL',            'https://sso.kneuralabs.com')
INTRANET_BASE_URL  = os.environ.get('INTRANET_BASE_URL',  'https://intranet.kneuralabs.com')
SSO_SHARED_SECRET  = os.environ.get('SSO_SHARED_SECRET',  '')
if not SSO_SHARED_SECRET:
    # Dev mode only (production raised above).
    warnings.warn(
        'SSO_SHARED_SECRET is not set — SSO callbacks cannot be verified. '
        'Set the SSO_SHARED_SECRET environment variable.',
        RuntimeWarning, stacklevel=1,
    )
    SSO_SHARED_SECRET = 'insecure-dev-' + os.urandom(32).hex()
_SSO_TOKEN_MAX_AGE = 300  # seconds

# Application-specific KDF salt for the XLSX encryption key.
# This is a domain-separation constant, NOT a password salt — it intentionally
# stays fixed so the same EXCEL_PASSWORD always produces the same Fernet key.
_KDF_SALT = b'kneura_net_v2'


def _fernet() -> Fernet:
    kdf = PBKDF2HMAC(
        algorithm=hashes.SHA256(),
        length=32,
        salt=_KDF_SALT,
        iterations=390_000,
    )
    key = base64.urlsafe_b64encode(kdf.derive(EXCEL_PASSWORD.encode()))
    return Fernet(key)


# ── In-memory workbook cache ───────────────────────────────────────────────────
# Avoids re-decrypting and re-parsing the XLSX on every request.
# Invalidated by mtime change whenever the file is written.
_wb_cache: dict = {'wb': None, 'mtime': 0.0}
_wb_lock  = Lock()   # guards both the cache dict and read-modify-write XLSX ops


def _load_wb() -> openpyxl.Workbook:
    """Return the cached workbook, re-loading from disk only when the file changed.
    Caller must NOT hold _wb_lock (this function acquires it internally)."""
    with _wb_lock:
        return _load_wb_locked()


def _load_wb_locked() -> openpyxl.Workbook:
    """Internal: load/return workbook. Caller MUST hold _wb_lock."""
    try:
        mtime = os.path.getmtime(DATA_FILE) if os.path.exists(DATA_FILE) else 0.0
    except OSError:
        mtime = 0.0

    if _wb_cache['wb'] is not None and _wb_cache['mtime'] == mtime:
        return _wb_cache['wb']

    if not os.path.exists(DATA_FILE) or os.path.getsize(DATA_FILE) == 0:
        wb = _blank_wb()
    else:
        try:
            with open(DATA_FILE, 'rb') as f:
                raw = f.read()
            wb = openpyxl.load_workbook(io.BytesIO(_fernet().decrypt(raw)))
        except Exception:
            wb = _blank_wb()

    _wb_cache.update({'wb': wb, 'mtime': mtime})
    return wb


def _blank_wb() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Users'
    ws.append(['EmployeeID', 'PasswordHash', 'Status', 'LastLogin', 'CreatedAt'])
    return wb


def _save_wb_locked(wb: openpyxl.Workbook) -> None:
    """Persist workbook to disk. Caller MUST hold _wb_lock."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    encrypted = _fernet().encrypt(buf.read())
    with open(DATA_FILE, 'wb') as f:
        f.write(encrypted)
    # Refresh cache so subsequent reads don't re-parse from disk
    try:
        mtime = os.path.getmtime(DATA_FILE)
    except OSError:
        mtime = 0.0
    _wb_cache.update({'wb': wb, 'mtime': mtime})
    # NOTE: git commit deliberately removed.
    # Persist DATA.xlsx through your CI/CD pipeline, not from inside the web process.


# ── User operations ────────────────────────────────────────────────────────────

def get_user(employee_id: str) -> dict | None:
    """Read-only lookup; uses cached workbook."""
    wb = _load_wb()
    for row in wb.active.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and str(row[0]).strip() == str(employee_id).strip():
            return {
                'id':            row[0],
                'password_hash': row[1],
                'status':        row[2],
                'last_login':    row[3],
                'created_at':    row[4],
            }
    return None


def upsert_user(employee_id: str, password_hash: str, status: str = 'active') -> None:
    """Insert or update a user row; holds lock for full read-modify-write."""
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with _wb_lock:
        wb = _load_wb_locked()
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if row[0].value is not None and str(row[0].value).strip() == str(employee_id).strip():
                row[1].value = password_hash
                row[2].value = status
                row[3].value = now
                _save_wb_locked(wb)
                return
        ws.append([str(employee_id), password_hash, status, now, now])
        _save_wb_locked(wb)


def update_last_login(employee_id: str) -> None:
    """Stamp last-login timestamp; holds lock for full read-modify-write."""
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with _wb_lock:
        wb = _load_wb_locked()
        for row in wb.active.iter_rows(min_row=2):
            if row[0].value is not None and str(row[0].value).strip() == str(employee_id).strip():
                row[3].value = now
                _save_wb_locked(wb)
                return


# ── Employee validation ────────────────────────────────────────────────────────

def check_employee_roster(employee_id: str) -> str:
    """Returns 'found', 'revoked', or 'not_found'."""
    try:
        resp = requests.get(EMPLOYEE_DATA_URL, timeout=10)
        resp.raise_for_status()
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        revoke_kw = {'revok', 'inactive', 'terminat', 'disabled', 'suspended'}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            if str(row[0]).strip() == str(employee_id).strip():
                for cell in row[1:]:
                    if cell and any(k in str(cell).lower() for k in revoke_kw):
                        return 'revoked'
                return 'found'
        return 'not_found'
    except Exception:
        return 'error'


# ── LinkedIn followers ─────────────────────────────────────────────────────────
# The public company page embeds the follower count in its meta description
# ("KneuraLabs | 1,234 followers on LinkedIn"). Fetched server-side because
# linkedin.com does not allow cross-origin browser requests, and cached so we
# hit LinkedIn at most once per TTL regardless of dashboard traffic.

LINKEDIN_COMPANY_URL = 'https://www.linkedin.com/company/kneuralabs/'
LINKEDIN_COMPANY_ID = '112376100'
# The pages-extensions follow-button endpoint is built for anonymous embedding
# and is far less likely to be authwalled than the company page itself.
_LI_SOURCE_URLS = (
    f'https://www.linkedin.com/pages-extensions/FollowCompany'
    f'?id={LINKEDIN_COMPANY_ID}&counter=bottom',
    LINKEDIN_COMPANY_URL,
)
_LI_CACHE_TTL = 6 * 3600  # seconds
_li_cache = {'followers': None, 'fetched_at': 0.0}
_li_lock = Lock()


def _parse_linkedin_followers(html: str) -> int | None:
    match = re.search(r'([\d][\d,.]*)\s+followers', html, re.IGNORECASE)
    if not match:
        # Follow-button markup carries a bare number in a follower-count element.
        match = re.search(r'follower[^>]*>\s*([\d][\d,.]*)\s*<', html,
                          re.IGNORECASE)
    if not match:
        return None
    return int(re.sub(r'\D', '', match.group(1)))


def _fetch_linkedin_followers() -> int | None:
    for url in _LI_SOURCE_URLS:
        try:
            resp = requests.get(
                url,
                timeout=10,
                headers={
                    'User-Agent': ('Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
                                   'AppleWebKit/537.36 (KHTML, like Gecko) '
                                   'Chrome/124.0 Safari/537.36'),
                    'Accept-Language': 'en-US,en;q=0.9',
                },
            )
            resp.raise_for_status()
        except requests.RequestException:
            continue
        followers = _parse_linkedin_followers(resp.text)
        if followers is not None:
            return followers
    return None


# ── SSO helpers ────────────────────────────────────────────────────────────────

def _sso_login_url() -> str:
    callback = f'{INTRANET_BASE_URL}/sso/callback'
    return f'{SSO_URL}/?redirect={callback}'


def _verify_sso_token(token: str) -> str:
    """Returns employee_id or raises BadSignature / SignatureExpired."""
    s = URLSafeTimedSerializer(SSO_SHARED_SECRET)
    data = s.loads(token, salt='sso-callback', max_age=_SSO_TOKEN_MAX_AGE)
    return data['employee_id']


# ── Auth decorator ─────────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'employee_id' not in session:
            return redirect(_sso_login_url())
        return f(*args, **kwargs)
    return decorated


# ── Routes ─────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    if 'employee_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(_sso_login_url())


@app.route('/login')
def login():
    return redirect(_sso_login_url())


@app.route('/dashboard')
@login_required
def dashboard():
    return render_template('dashboard.html', employee_id=session['employee_id'])


@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    employee_id = session['employee_id']
    first_login = session.get('first_login', False)

    if request.method == 'POST':
        current_pw  = request.form.get('current_password', '').strip()
        new_pw      = request.form.get('new_password',      '').strip()
        confirm_pw  = request.form.get('confirm_password',  '').strip()

        if not current_pw or not new_pw or not confirm_pw:
            flash('All fields are required.', 'error')
            return render_template('change_password.html', first_login=first_login)

        if new_pw != confirm_pw:
            flash('New passwords do not match.', 'error')
            return render_template('change_password.html', first_login=first_login)

        if len(new_pw) < 8:
            flash('Password must be at least 8 characters.', 'error')
            return render_template('change_password.html', first_login=first_login)

        if not any(c.isalpha() for c in new_pw) or not any(c.isdigit() for c in new_pw):
            flash('Password must contain at least one letter and one digit.', 'error')
            return render_template('change_password.html', first_login=first_login)

        user = get_user(employee_id)
        if not user:
            flash('User not found.', 'error')
            return render_template('change_password.html', first_login=first_login)

        if not bcrypt.checkpw(current_pw.encode(), user['password_hash'].encode()):
            flash('Current password is incorrect.', 'error')
            return render_template('change_password.html', first_login=first_login)

        # Deny-list: the configured onboarding password and the well-known
        # legacy default (kept here purely as a denied value, not a credential).
        if new_pw == DEFAULT_PASSWORD or new_pw == 'Kneuralabs@2026':
            flash('You cannot reuse the default password.', 'error')
            return render_template('change_password.html', first_login=first_login)

        new_hash = bcrypt.hashpw(new_pw.encode(), bcrypt.gensalt(rounds=12)).decode()
        upsert_user(employee_id, new_hash)
        session['first_login'] = False
        flash('Password changed successfully.', 'success')
        return redirect(url_for('dashboard'))

    return render_template('change_password.html', first_login=first_login)


@app.route('/api/linkedin-followers')
@login_required
def linkedin_followers():
    now = time.time()
    with _li_lock:
        cached = _li_cache['followers']
        fresh = cached is not None and (now - _li_cache['fetched_at']) < _LI_CACHE_TTL
    if fresh:
        return jsonify({'followers': cached})

    try:
        count = _fetch_linkedin_followers()
    except Exception:
        count = None

    if count is not None:
        with _li_lock:
            _li_cache.update({'followers': count, 'fetched_at': now})
        return jsonify({'followers': count})

    # Fetch failed — serve the stale value if we have one rather than nothing.
    if cached is not None:
        return jsonify({'followers': cached, 'stale': True})
    return jsonify({'followers': None}), 503


@app.route('/sso/callback')
def sso_callback():
    token = request.args.get('token', '')
    if not token:
        flash('SSO authentication failed. Please try again.', 'error')
        return redirect(_sso_login_url())

    try:
        employee_id = _verify_sso_token(token)
    except SignatureExpired:
        flash('SSO session expired. Please sign in again.', 'error')
        return redirect(_sso_login_url())
    except BadSignature:
        flash('Invalid SSO token. Please sign in again.', 'error')
        return redirect(_sso_login_url())

    # Rotate the session on login to prevent session fixation.
    session.clear()
    session['employee_id'] = employee_id
    user = get_user(employee_id)

    if user is None:
        hashed = bcrypt.hashpw(DEFAULT_PASSWORD.encode(), bcrypt.gensalt(rounds=12)).decode()
        upsert_user(employee_id, hashed)
        session['first_login'] = True
        flash('Welcome! Please change your default password.', 'info')
        return redirect(url_for('change_password'))

    session['first_login'] = False
    update_last_login(employee_id)
    return redirect(url_for('dashboard'))


@app.route('/logout')
def logout():
    session.clear()
    return redirect(_sso_login_url())


if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=5000)
